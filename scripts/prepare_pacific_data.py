#!/usr/bin/env python3
"""
Preprocessing für das Dashboard "Pacific Cyclones & Their Impacts".

Erzeugt schlanke Tidy-Tabellen aus den beiden brauchbaren Rohdatensätzen im
Ordner "Recherche Data":

  1. EM-DAT (public_emdat_*.xlsx)  -> zyklon-spezifische Impacts (Tote, Betroffene, Schäden)
  2. UN WPP 2024 (WPP2024_GEN_F01_*.xlsx) -> Bevölkerung pro Land/Jahr (zum Normalisieren)

Beide werden auf die pazifischen Inselstaaten/-territorien (PICTs) gefiltert und
in das gemeinsame View-Schema `country, year, ...` überführt
(siehe docs/08_Dashboard-Konzept.md §5).

Outputs (-> Data/processed/):
  - emdat_pacific_storms_events.csv        (1 Zeile je Sturm-Event, inkl. Lat/Lon für Karte)
  - emdat_pacific_storms_by_country_year.csv (aggregiert je Land/Jahr -> View 2/3)
  - wpp_pacific_population.csv              (Bevölkerung je Land/Jahr -> Normalisierung/Bubble-Größe)

Aufruf:  python3 scripts/prepare_pacific_data.py
"""

from __future__ import annotations

import csv
from collections import defaultdict
from pathlib import Path

import openpyxl

# --- Pfade ---------------------------------------------------------------
ROOT = Path(__file__).resolve().parent.parent
RAW = ROOT / "Recherche Data"
OUT = ROOT / "Data" / "processed"
OUT.mkdir(parents=True, exist_ok=True)

EMDAT_XLSX = RAW / "public_emdat_2026-06-15.xlsx"
WPP_XLSX = RAW / "WPP2024_GEN_F01_DEMOGRAPHIC_INDICATORS_FULL.xlsx"

# --- Pazifische Inselstaaten/-territorien (PICTs), ISO3 ------------------
# 21 PICTs des Pacific Data Hub + American Samoa
PACIFIC_ISO = {
    "ASM": "American Samoa",
    "COK": "Cook Islands",
    "FJI": "Fiji",
    "PYF": "French Polynesia",
    "GUM": "Guam",
    "KIR": "Kiribati",
    "MHL": "Marshall Islands",
    "FSM": "Micronesia (Fed. States of)",
    "NRU": "Nauru",
    "NCL": "New Caledonia",
    "NIU": "Niue",
    "MNP": "Northern Mariana Islands",
    "PLW": "Palau",
    "PNG": "Papua New Guinea",
    "PCN": "Pitcairn",
    "WSM": "Samoa",
    "SLB": "Solomon Islands",
    "TKL": "Tokelau",
    "TON": "Tonga",
    "TUV": "Tuvalu",
    "VUT": "Vanuatu",
    "WLF": "Wallis and Futuna Islands",
}


def _num(v):
    """xlsx-Zelle -> float|None (leere/Platzhalter werden None)."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ("", "...", "-", "NA", "N/A"):
        return None
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return None


def _int(v):
    f = _num(v)
    return int(f) if f is not None else None


# --- 1) EM-DAT: Pazifik-Stürme ------------------------------------------
def process_emdat() -> None:
    print(f"[EM-DAT] lese {EMDAT_XLSX.name} ...")
    # Hinweis: read_only=True meldet bei dieser Datei eine falsche Dimension
    # (nur Spalte A) -> bewusst read_only=False, dann über die echte Zellmatrix iterieren.
    wb = openpyxl.load_workbook(EMDAT_XLSX, read_only=False, data_only=True)
    ws = wb["EM-DAT Data"]

    H = {ws.cell(1, j).value: j for j in range(1, ws.max_column + 1)}  # 1-basierte Spaltenindizes

    def g(row_idx, col):
        j = H.get(col)
        return ws.cell(row_idx, j).value if j else None

    events = []
    for row in range(2, ws.max_row + 1):
        iso = g(row, "ISO")
        if iso not in PACIFIC_ISO:
            continue
        if g(row, "Disaster Type") != "Storm":
            continue  # nur Stürme/tropische Zyklone
        events.append(
            {
                "disno": g(row, "DisNo."),
                "iso3": iso,
                "country": PACIFIC_ISO[iso],
                "year": _int(g(row, "Start Year")),
                "month": _int(g(row, "Start Month")),
                "subtype": g(row, "Disaster Subtype"),
                "event_name": g(row, "Event Name"),
                "lat": _num(g(row, "Latitude")),
                "lon": _num(g(row, "Longitude")),
                "magnitude": _num(g(row, "Magnitude")),
                "magnitude_scale": g(row, "Magnitude Scale"),
                "total_deaths": _int(g(row, "Total Deaths")),
                "total_affected": _int(g(row, "Total Affected")),
                "total_damage_kusd": _num(g(row, "Total Damage ('000 US$)")),
                "total_damage_adj_kusd": _num(g(row, "Total Damage, Adjusted ('000 US$)")),
            }
        )

    events.sort(key=lambda e: (e["country"], e["year"] or 0, e["month"] or 0))

    # 1a) Event-Level CSV (für Karte / View 1-Kreuzbezug)
    ev_cols = [
        "disno", "iso3", "country", "year", "month", "subtype", "event_name",
        "lat", "lon", "magnitude", "magnitude_scale",
        "total_deaths", "total_affected", "total_damage_kusd", "total_damage_adj_kusd",
    ]
    ev_path = OUT / "emdat_pacific_storms_events.csv"
    with ev_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=ev_cols)
        w.writeheader()
        w.writerows(events)
    print(f"[EM-DAT] {len(events)} Sturm-Events -> {ev_path.relative_to(ROOT)}")

    # 1b) Aggregat je Land/Jahr (gemeinsames View-Schema)
    agg = defaultdict(lambda: {"events": 0, "deaths": 0, "affected": 0, "damage_kusd": 0.0})
    for e in events:
        if e["year"] is None:
            continue
        k = (e["iso3"], e["country"], e["year"])
        a = agg[k]
        a["events"] += 1
        a["deaths"] += e["total_deaths"] or 0
        a["affected"] += e["total_affected"] or 0
        a["damage_kusd"] += e["total_damage_kusd"] or 0.0

    by_path = OUT / "emdat_pacific_storms_by_country_year.csv"
    with by_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["iso3", "country", "year", "storm_events", "deaths", "affected", "damage_kusd"])
        for (iso, country, year), a in sorted(agg.items(), key=lambda kv: (kv[0][1], kv[0][2])):
            w.writerow([iso, country, year, a["events"], a["deaths"], a["affected"], round(a["damage_kusd"], 1)])
    print(f"[EM-DAT] {len(agg)} Land/Jahr-Zeilen -> {by_path.relative_to(ROOT)}")


# --- 2) UN WPP 2024: Bevölkerung Pazifik --------------------------------
def process_wpp() -> None:
    print(f"[WPP]   lese {WPP_XLSX.name} (groß, bitte kurz warten) ...")
    wb = openpyxl.load_workbook(WPP_XLSX, read_only=True, data_only=True)
    ws = wb["Estimates"]

    ISO_COL = "ISO3 Alpha-code"
    TYPE_COL = "Type"
    YEAR_COL = "Year"
    POP_COL = "Total Population, as of 1 July (thousands)"

    H = None
    out = []
    for row in ws.iter_rows(values_only=True):
        if H is None:
            if row and ISO_COL in row:
                H = {name: i for i, name in enumerate(row)}
            continue
        iso = row[H[ISO_COL]]
        if iso not in PACIFIC_ISO:
            continue
        if row[H[TYPE_COL]] != "Country/Area":
            continue
        pop = _num(row[H[POP_COL]])
        out.append(
            {
                "iso3": iso,
                "country": PACIFIC_ISO[iso],
                "year": _int(row[H[YEAR_COL]]),
                "population": int(round(pop * 1000)) if pop is not None else None,  # Köpfe (thousands -> absolut)
            }
        )

    out.sort(key=lambda r: (r["country"], r["year"] or 0))
    wpp_path = OUT / "wpp_pacific_population.csv"
    with wpp_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["iso3", "country", "year", "population"])
        w.writeheader()
        w.writerows(out)
    countries = len({r["iso3"] for r in out})
    print(f"[WPP]   {len(out)} Zeilen ({countries} Länder) -> {wpp_path.relative_to(ROOT)}")


if __name__ == "__main__":
    # Rohquellen können nach der Bereinigung fehlen -> dann überspringen
    # (die fertigen Subsets in Data/processed/ bleiben erhalten).
    if EMDAT_XLSX.exists():
        process_emdat()
    else:
        print(f"[EM-DAT] übersprungen – Rohdatei fehlt ({EMDAT_XLSX.name}); ggf. neu beziehen.")
    if WPP_XLSX.exists():
        process_wpp()
    else:
        print(f"[WPP]   übersprungen – Rohdatei fehlt ({WPP_XLSX.name}); ggf. neu beziehen.")
    print("Fertig. Outputs in Data/processed/")
